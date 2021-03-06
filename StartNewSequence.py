from openpyxl import Workbook
from openpyxl import load_workbook
import time
import threading
import numpy as np
import xlsxwriter
import SwerleinFreq
import xlwings as xw
import wx
import winsound
import pyttsx
import sys
import win32com.client as win32

completedStartNewSequence=0
#-----Instruments used in Watt Bridge Software-----#
HP3458A_V=0
Ag53230A_V=0
FLUKE_V=0
rd31=0
HP3478A_V=0
RS232_6_WB=0
#-----Global Variables used by the Watt Bridge Software-----#
flukeError=0
FlukeErrorNumber=0
SetVoltsCell=0
SetAmpsCell=0
SetPhaseCell=0
SetVoltsCell=0
SetFrequencyCell=0
ReadingNumber=0
RowNumber=0
NumberOfReadings=0
DividerRange=0
Shunt=0
CTRatio=0
HEGFreq=0
SourceType=0
CHType=0
CalculationResults=0
RDPhase=0
SampleData=[]
WCount=0
WSign=0
VCount=0
VSign=0
Finished=0
ACVoltsRms=0
UncalFreqy=0
FFTVolts=0
FFTPhase=0
VRangeHigh=0
DCVoltageOffset=0
DCCurrentOffset=0
HighCurrentRange=0
IRangeLow=0
IRangeHigh=0
Chanel=0
SetVoltsPhase=0
SampleTime=0
input1=[]
input2=[]
input3=[]
inputTotal=[]
chanel=0
VRangeHigh=0
FFT = []
#---------------------------------------------------#
#Temporary variables. To be checked later
DCVRange=0
ActiveRow=0
lineCurrent = 0
lineVolts = 0
phase = 0

def immediateShutDown():
    '''Turns off the power supply for emergency purposes. Such as if an error were to occur.'''
    FLUKE_V.write("OUTP:STAT OFF")

def emailMessage(subject, message):
    '''Send an email to recipients if any error were to occur or when the data sequence has finished its measurements.'''
    outlook = win32.Dispatch('Outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = 'tom.stewart@callaghaninnovation.govt.nz'
    mail.CC = 'ikram.singh@callaghaninnovation.govt.nz'
    mail.Subject = subject
    mail.Body = message+' \n\n\n\n\nKind Regards, \nWatt Bridge Computer \nEmail generated using Python'
    mail.Send()

def textToVoice(wattBridgeGUI,text):
    '''Computer speaks the stage that the software is currently in. Can be enable/disabled by the user in GUI.'''
    if wattBridgeGUI.TextToVoice.GetValue()==True:
        winsound.Beep(40,750)
        engine = pyttsx.init()
        engine.setProperty('rate',121)
        engine.say(text)
        engine.runAndWait()

def getExcelColumn(column):
    '''Returns the character(s) for the excel column number (i.e. Converts column Integer to Alpha)'''
    columnCharacter = str(xlsxwriter.utility.xl_col_to_name(column-1))
    return columnCharacter

def setupChanel(wattBridgeGUI):
    '''Sends various commands to the FLUKE power supply. Checks to see if the fittings are placed properly.
    Sets the FLUKE power supply to the appropriate DC offset and phase values.'''
    start = time.clock()
    FLUKE_V.write("SOUR:PHAS" +str(Chanel)+":FITT?") #Output to FLUKE_V with "SOUR:PHAS" , Chanel , ":FITT?", term.=LF
    time.sleep(0.5) #Delay for 0.5 seconds
    #Enter from FLUKE_V up to 256 bytes, stop on EOS=LF
    #Store in Phase On from FLUKE_V
    PhaseOn = FLUKE_V.read()
    if PhaseOn==0:
        wattBridgeGUI.WattBridgeEventsLog.AppendText("Cause error: One of the phases you have tried turn on is not fitted \n") #Update event log.
        emailMessage('Watt Bridge Error', 'Cause error: One of the phases you have tried turn on is not fitted')
        immediateShutDown()
        sys.exit()
        time.sleep(1)
    if SourceType == "FLUHIGH":
        FLUKE_V.write("SOUR:PHAS" +str(Chanel)+ ":CURR:EAMP:FITT?") #Output to FLUKE_V with "SOUR:PHAS" , Chanel , ":CURR:EAMP:FITT?", term.=LF
        #Enter from FLUKE_V up to 256 bytes, stop on EOS=LF
        #Store in Amp Fitted from FLUKE_V
        AmpFitted = FLUKE_V.read()
        if AmpFitted==0:
            wattBridgeGUI.WattBridgeEventsLog.AppendText("Cause error: A 52120A unit is not fitted to the phase you have selected \n") #Update event log.
            emailMessage('Watt Bridge Error', "Cause error: A 52120A unit is not fitted to the phase you have selected \n")
            immediateShutDown()
            sys.exit()
        if wattBridgeGUI.OutputAutoHigh.GetCurrentSelection()==0:
            FLUKE_V.write("SOUR:PHAS" +str(Chanel)+ ":CURR:EAMP:TERM:MODE AUTO") #Output to FLUKE_V with "SOUR:PHAS" , Chanel , ":CURR:EAMP:TERM:MODE AUTO", term.=LF
            print("AUTO")
        elif wattBridgeGUI.OutputAutoHigh.GetCurrentSelection()==1:
            FLUKE_V.write("SOUR:PHAS" +str(Chanel)+ ":CURR:EAMP:TERM:MODE HIGH") #Output to FLUKE_V with "SOUR:PHAS" , Chanel , ":CURR:EAMP:TERM:MODE HIGH", term.=LF
            print("HIGH")
        FLUKE_V.write("SOUR:PHAS" +str(Chanel)+ ":CURR:EAMP:RANG " +str(0)+  "," +str(HighCurrentRange)) #Output to FLUKE_V with "SOUR:PHAS" , Chanel , ":CURR:EAMP:RANG " , 0 , "," , High Current Range, term.=LF  
    else:
        if SetAmpsCell>21:
            wattBridgeGUI.WattBridgeEventsLog.AppendText("Cause error: You have selected a current value above 21A, please use the FLUHIGH source in Excel and retry \n") #Update event log.
            emailMessage('Watt Bridge Error', "Cause error: You have selected a current value above 21A, please use the FLUHIGH source in Excel and retry \n")
            immediateShutDown()
            sys.exit()
        FLUKE_V.write("SOUR:PHAS" +str(Chanel)+ ":CURR:RANG " +str(IRangeLow)+ "," +str(IRangeHigh)) #Output to FLUKE_V with "SOUR:PHAS" , Chanel , ":CURR:RANG " , I RangeLow , "," , I RangeHigh, term.=LF
    FLUKE_V.write("UNIT:MHAR:CURR ABS") #Output to FLUKE_V with "UNIT:MHAR:CURR ABS", term.=LF
    FLUKE_V.write("UNIT:MHAR:VOLT ABS") #Output to FLUKE_V with "UNIT:MHAR:VOLT ABS", term.=LF
    FLUKE_V.write("SOUR:PHAS" +str(Chanel) + ":VOLT:RANG " + "0," +str(VRangeHigh)) #Output to FLUKE_V with "SOUR:PHAS" , Chanel , ":VOLT:RANG " , "0," , V RangeHigh , term.=LF
    FLUKE_V.write("SOUR:PHAS" +str(Chanel)+ ":CURR:MHAR:STAT ON") #Output to FLUKE_V with "SOUR:PHAS" , Chanel , ":CURR:MHAR:STAT ON", term.=LF
    FLUKE_V.write("SOUR:PHAS" +str(Chanel)+ ":CURR:MHAR:HARM1 " +str(SetAmpsCell) + "," +str(SetPhaseCell)) #Output to FLUKE_V with "SOUR:PHAS" , Chanel , ":CURR:MHAR:HARM1 " , Set amps cell , "," , Set phase cell, term.=LF
    FLUKE_V.write("SOUR:PHAS" +str(Chanel)+ ":CURR:MHAR:HARM0 " +str(DCCurrentOffset)+ "," +str(0)) #Output to FLUKE_V with "SOUR:PHAS" , Chanel , ":CURR:MHAR:HARM0 " , DC Current Offset , "," , 0, term.=LF
    FLUKE_V.write("SOUR:PHAS" +str(Chanel)+ ":VOLT:MHAR:STAT ON") #Output to FLUKE_V with "SOUR:PHAS" , Chanel , ":VOLT:MHAR:STAT ON", term.=LF
    FLUKE_V.write("SOUR:PHAS" +str(Chanel)+ ":VOLT:MHAR:HARM1 " +str(SetVoltsCell)+ "," +str(SetVoltsPhase)) #Output to FLUKE_V with "SOUR:PHAS" , Chanel , ":VOLT:MHAR:HARM1 " , Set volts cell , "," , Set Volts Phase, term.=LF
    FLUKE_V.write("SOUR:PHAS" +str(Chanel)+ ":VOLT:MHAR:HARM0 " +str(DCVoltageOffset)+ "," +str(0))#Output to FLUKE_V with "SOUR:PHAS" , Chanel , ":VOLT:MHAR:HARM0 " , DC Voltage Offset , "," , 0, term.=LF
    FLUKE_V.write("SOUR:PHAS" +str(Chanel)+ ":VOLT:STAT " + "ON") #Output to FLUKE_V with "SOUR:PHAS" , Chanel , ":VOLT:STAT " , "ON", term.=LF
    FLUKE_V.write("SOUR:PHAS" +str(Chanel)+ ":CURR:STAT " + "ON") #Output to FLUKE_V with "SOUR:PHAS" , Chanel , ":CURR:STAT " , "ON", term.=LF
    final = time.clock()
    print('SetupChannel time: '+ str(final-start))
def setPhases():
    '''Sets the phase value.'''
    global SetPhaseCell
    if abs(SetPhaseCell) > 180.0:
        if SetPhaseCell>0:
            Sum = SetPhaseCell-360
            SetPhaseCell = Sum
        else:
            Sum = SetPhaseCell+360
            SetPhaseCell = Sum
def powerFluke(wattBridgeGUI,ws):
    '''Checks to see if the settings being sent to the FLUKE power supply are suitable. Selects
    values for the DC offsets and phase values for the appropriate input ranges.'''
    start = time.clock()
    global SetVoltsCell,SetPhaseCell,SetAmpsCell,SetFrequencyCell,VRangeHigh,DCVoltageOffset
    global HighCurrentRange,DCCurrentOffset,IRangeLow,IRangeHigh,Chanel,SetVoltsPhase,FlukeErrorNumber
    global flukeError,SetVoltsPhase 
    FLUKE_V.write("*CLS") #Output to FLUKE_V with "*CLS", term.=LF
    FLUKE_V.write("*RST") #Output to FLUKE_V with "*RST", term.=LF
    FLUKE_V.write("OUTP:SENS 0") #Output to FLUKE_V with "OUTP:SENS 0", term.=LF
    if wattBridgeGUI.Flukeramp.GetValue()<2:
        wattBridgeGUI.WattBridgeEventsLog.AppendText("Error message: Ramp time less than 2 seconds \n") #Update event log.
        emailMessage('Watt Bridge Error', "Error message: Ramp time less than 2 seconds \n")
        immediateShutDown()
        sys.exit()
    FLUKE_V.write("OUTP:RAMP:TIME "+str(wattBridgeGUI.Flukeramp.GetValue())) #Output to FLUKE_V with "OUTP:RAMP:TIME " , Fluke Ramp (s), term.=LF
    SetVoltsCell = ws['D'+str(ActiveRow)].value #Obtain set voltage value from Excel Sheet
    SetPhaseCell = ws['E'+str(ActiveRow)].value #Obtain set phase value from Excel Sheet
    setPhases() #Execute Set Phases function
    SetAmpsCell = ws['C'+str(ActiveRow)].value #Obtain set amps value from Excel Sheet
    SetFrequencyCell = ws['I'+str(ActiveRow)].value #Obtain set frequency value from Excel Sheet
    if SetVoltsCell>1008:
        wattBridgeGUI.WattBridgeEventsLog.AppendText("Cause error: Voltage value out of range \n") #Update event log.
        emailMessage('Watt Bridge Error', "Cause error: Voltage value out of range \n")
        immediateShutDown()
        sys.exit()
    if SetVoltsCell>250:
        wattBridgeGUI.WattBridgeEventsLog.AppendText("Cause error: Voltage selected is above 250V \n") #Update event log.
        emailMessage('Watt Bridge Error', "Cause error: Voltage selected is above 250V \n")
        immediateShutDown()
        sys.exit()
    if SetAmpsCell>120:
        wattBridgeGUI.WattBridgeEventsLog.AppendText("Cause error: Current value out of range \n") #Update event log.
        emailMessage('Watt Bridge Error', "Cause error: Current value out of range \n")
        immediateShutDown()
        sys.exit()
    #Specific cases for SetVoltsCell
    if 0 <= SetVoltsCell <= 22.9:
        VRangeHigh=23
        DCVoltageOffset=-0.00094
    elif 22.9 <= SetVoltsCell <= 44.9:
        VRangeHigh=45
        DCVoltageOffset=-0.0018
    elif 44.9 <= SetVoltsCell <= 89.9:
        VRangeHigh=90
        DCVoltageOffset=-0.00673
    elif 89.9 <= SetVoltsCell <= 179.9:
        VRangeHigh=180
        DCVoltageOffset=-0.01299
    elif 179.9 <= SetVoltsCell <= 359.9:
        VRangeHigh=360
        DCVoltageOffset=-0.02054
    elif 359.9 <= SetVoltsCell <= 649.9:
        VRangeHigh=650
        DCVoltageOffset=-0.02558
    elif 649.9 <= SetVoltsCell <= 1008:
        VRangeHigh=1008
        DCVoltageOffset=-0.0264
    if SourceType=="FLUHIGH":
        if 0 <= SetAmpsCell <= 1.999:
            HighCurrentRange=2
            DCCurrentOffset=-0.0001
        elif 1.999 <= SetAmpsCell <= 19.99:
            HighCurrentRange=20
            DCCurrentOffset=-0.001027
        elif 19.99 <= SetAmpsCell <= 120:
            HighCurrentRange=120
            DCCurrentOffset=0.000254
    else:
        if 0 <= SetAmpsCell <= 0.249:
            IRangeLow=0.05
            IRangeHigh=0.25
            DCCurrentOffset=1.3e-05
        elif 0.249 <= SetAmpsCell <= 0.499:
            IRangeLow=0.05
            IRangeHigh=0.5
            DCCurrentOffset=3.6e-05
        elif 0.499 <= SetAmpsCell <= 0.999:
            IRangeLow=0.1
            IRangeHigh=1
            DCCurrentOffset=6.8e-05
        elif 0.999 <= SetAmpsCell <= 1.999:
            IRangeLow=0.2
            IRangeHigh=2
            DCCurrentOffset=0.000133
        elif 1.999 <= SetAmpsCell <= 4.99:
            IRangeLow=0.5
            IRangeHigh=5
            DCCurrentOffset=0.000345
        elif 4.99 <= SetAmpsCell <= 9.99:
            IRangeLow=1
            IRangeHigh=10
            DCCurrentOffset=0.000786
        elif 9.99 <= SetAmpsCell <= 21:
            IRangeLow=2
            IRangeHigh=21
            DCCurrentOffset=0.001675
    if wattBridgeGUI.SourceCh1.GetValue()==True:
        Chanel = 1
        SetPhaseCell=SetPhaseCell
        SetVoltsPhase = 0
        setupChanel(wattBridgeGUI) #Execute Setup Chanel function
    if wattBridgeGUI.SourceCh2.GetValue()==True:
        Chanel = 2
        SetPhaseCell=SetPhaseCell
        SetVoltsPhase = -120
        setupChanel(wattBridgeGUI) #Execute Setup Chanel function 
    if wattBridgeGUI.SourceCh3.GetValue()==True:
        Chanel = 3
        SetPhaseCell=SetPhaseCell
        SetVoltsPhase = 120
        setupChanel(wattBridgeGUI) #Execute Setup Chanel function 
    FlukeErrorNumber=1
    while FlukeErrorNumber!=0:
        FLUKE_V.write("SOUR:FREQ "+str(SetFrequencyCell)) #Output to FLUKE_V with "SOUR:FREQ " , Set frequency cell, term.=LF
        FLUKE_V.write("SYST:ERR?") #Output to FLUKE_V with "SYST:ERR?", term.=LF
        #Enter from FLUKE_V(1) up to 512 bytes, stop on EOS=LF
        flukeError = str(FLUKE_V.read()) #Store in Fluke Error from FLUKE_V
        print("flukeError: "+str(flukeError))
        time.sleep(0.5) #Delay for 0.5 seconds
        #Calculate Vector Index with v=Fluke Error i=0
        FlukeErrorNumber = int(flukeError[0]) #Store in Fluke Error number from Vector Index
        print("FlukeErrorNumber: "+str(FlukeErrorNumber))
        if FlukeErrorNumber!=0: #If/Then Fluke error with x=Fluke Error number
            VectorIndex = flukeError[1]#Calculate Vector Index with v=Fluke Error i=1
            wattBridgeGUI.WattBridgeEventsLog.AppendText("Cause error: " + str(VectorIndex)+ "\n") #Update event log. #Cause error General Error code=20010, text=Vector Index
            emailMessage('Watt Bridge Error', "Cause error: " + str(VectorIndex)+ "\n")
            immediateShutDown()
            sys.exit()
    FLUKE_V.write("OUTP:STAT ON")#Output to FLUKE_V with "OUTP:STAT ON", term.=LF
    time.sleep(float(wattBridgeGUI.Flukeramp.GetValue())) #Delay for Fluke Ramp (s) seconds
    FLUKE_V.write("OUTP:STAT?")#Output to FLUKE_V with "OUTP:STAT?", term.=LF
    #Enter from FLUKE_V up to 256 bytes, stop on EOS=LF
    isFlukeOff = FLUKE_V.read() #Store in Is Fluke Off from FLUKE_V
    time.sleep(0.5) #Delay for 0.5 seconds
    if isFlukeOff < 0.5: #If/Then Fluke not On with x=Is Fluke Off
        wattBridgeGUI.WattBridgeEventsLog.AppendText("Cause error: Power output unsuccessful. Check connections. \n") #Update event log. 
        emailMessage('Watt Bridge Error', "Cause error: Power output unsuccessful. Check connections. \n")
        immediateShutDown()
        sys.exit()
    time.sleep(12) #Delay for 12 seconds
    final = time.clock()
    print('PowerFluke time: '+ str(final-start))
## def powerCH5500(ws):
##     global CHType
##     CHType = ws['B7'].value #Obtain CHType
##     #Clear CH5500_V
##     if CHType<99:
##         #Clear CH5050_V
##         #Output to CH5050_V with "S" , "I", term.=LF
##         print("CHType < 99")
##     #Output to CH5500_V with "S", term.=LF
##     SetVoltsCell = ws['D'+str(ActiveRow)].value #Obtain set voltage value from Excel Sheet
##     SetVolts=0
##     if 9 <CHType< 12:
##         SetVolts = SetVoltsCell/2.497
##         #Output to CH5500_V with "O" , 0, term.=LF
##     else:
##         SetVolts = SetVoltsCell/1
##         #Output to CH5500_V with "O" , 180, term.=LF
##     #Output to CH5500_V with "R" , Set volts, term.=LF
##     SetAmpsCell = ws['C'+str(ActiveRow)].value #Obtain set amps value from Excel Sheet
##     if SetAmpsCell<0:
##         #Output to CH5500_V with "O" , 0, term.=LF
##         print('SetAmpsCell<0')
##     Absolutevalue = abs(SetAmpsCell)
##     #Output to CH5500_V with "V" , Absolute value, term.=LF
##     SetPhaseCell = ws['E'+str(ActiveRow)].value #Obtain set phase value from Excel Sheet
##     #Output to CH5500_V with "P" , SetPhaseCell, term.=LF
##     SetFrequencyCell = ws['I'+str(ActiveRow)].value #Obtain set frequency value from Excel Sheet
##     #Output to CH5500_V with "F" , SetFrequencyCell, term.=LF
##     #Output to CH5500_V with "N", term.=LF
##     if CHType<99:
##         #Output to CH5050_V with "N", term.=LF
##         #Output to CH5500_V with "N", term.=LF
##         #Output to CH5050_V with "O", term.=LF
##         print('CHType<99')
##     time.sleep(60) #Delay for 60 seconds

def setUpFFTVoltsAndPhase(ws):
    '''Calculates the Fast Fourier Transform (both magnitude and phase vectors) of the data from the 3458A.
    The FFT voltage and phase are obtained from from these vectors.'''
    global SampleData,FFTVolts,FFTPhase,FFT
    start = time.clock()
    SampleData=[] #Clear sample data
    time.sleep(2) #Delay for 2 seconds
    for FFTLoop in range(256):
        output = float(HP3458A_V.read()) #Enter from HP3458A_V up to 256 bytes, stop on EOS=LF
        SampleData.append(output) #Append to Sample Data from HP3458A_V
    FFT = np.fft.fft(SampleData) #Calculate FFT with freq=FFT freqy wave=Sample Data
    MagnitudeVector = np.abs(FFT) #Calculate MagnitudeVector with spectrum=FFT
    PhaseVector = np.angle(FFT) #Calculate PhaseVector with spectrum=FFT
    FFTVolts = MagnitudeVector[9]*2.0/(256.0*np.sqrt(2)) #Calculate FFT Volts with n=9 V=MagnitudeVector
    FFTPhase = PhaseVector[9]*180.0/(np.pi) #Calculate FFT Phase with n=9 V=PhaseVector
    final = time.clock()
    print('setupVoltsAndPhase time: '+ str(final-start))
    time.sleep(2)
def setUpFFT():
    '''Calculates the SampleTime which is dependent on the frequency obtain through Swerleins Algorithm'''
    global SampleTime, UncalFreqy
    start = time.clock()
    time.sleep(2) #For settling purposes
    UncalFreqy = SwerleinFreq.FNFreq() #Obtain the Frequency from 3458A
    SampleTime = 9/(256*UncalFreqy)
    print("SampleTime: "+str(SampleTime))
    HP3458A_V.write("preset fast;mem fifo;mformat sint;oformat ascii") #Output to HP3458A_V with "preset fast" , ";mem fifo" , ";mformat sint" , ";oformat ascii", term.=LF
    HP3458A_V.write("ssdc ;range 10;ssrc ext") #Output to HP3458A_V with "ssdc " , ";range 10" , ";ssrc ext", term.=LF
    HP3458A_V.write(";delay 1e-03;sweep "+str(SampleTime)+" , 256") #Output to HP3458A_V with ";delay 1e-03" , ";sweep " , Sample Time , "," , 256, term.=LF
    final = time.clock()
    print('SetupChannel: '+ str(final-start))
    time.sleep(2)
def findDialSettings(wattBridgeGUI,ws):
    '''Obtains the Dial setting from the Excel sheet and executes the setUpFFT 
    and setUpFFTVoltsAndPhase functions. The GUI gets updated at the same time.'''
    global WCount,VCount,WSign,VSign,UncalFreqy,lineCurrent, lineVolts, phase
    start = time.clock()
    ws['AC'+str(ActiveRow)].value = 0 #Set W dial cell to 0
    ws['AD'+str(ActiveRow)].value = "WP-" #Set W sign cell to "WP-"
    ws['AE'+str(ActiveRow)].value = 0 #Set V dial cell to 0
    ws['AF'+str(ActiveRow)].value = "VP-" #Set V sign cell to "VP-"
    ws['F'+str(ActiveRow)].value = DividerRange
    ws['G'+str(ActiveRow)].value = Shunt
    ws['H'+str(ActiveRow)].value = CTRatio
    ws['I'+str(ActiveRow)].value = HEGFreq
    ws['AM7'].value = "Min"
    UncalFreqy = SwerleinFreq.FNFreq() #Obtain the Frequency from 3458A. Used to be "Execute Frequency function".
    updateGUI(wattBridgeGUI,ws)
    ws['AB'+str(ActiveRow)].value=UncalFreqy #Set the exact frequency value in Excel sheet.
    setUpFFT() #Execute Set Up FFT function
    setUpFFTVoltsAndPhase(ws) #Execute FFT Volts & Phase function
    if FFTVolts<0.7:
        wattBridgeGUI.WattBridgeEventsLog.AppendText("Cause error: Source Voltage Error \n") #Update event log.
        emailMessage('Watt Bridge Error', "Cause error: Source Voltage Error \n")
        immediateShutDown()
        sys.exit()
    ws[getExcelColumn(36+7*(ReadingNumber-1))+str(ActiveRow)].value=FFTVolts #Set the FFT ref volts value in Excel sheet.
    ws[getExcelColumn(37+7*(ReadingNumber-1))+str(ActiveRow)].value=FFTPhase #Set the FFT ref phase value in Excel sheet.
    #Output to RS232 6 WB with "DD", term.=CR, wait for completion?=1
    #Close RS232 6 WB
    RS232_6_WB.write("DD\r")
    time.sleep(1)
    setUpFFTVoltsAndPhase(ws) #Execute FFT Volts & Phase function
    ws[getExcelColumn(33+7*(ReadingNumber-1))+str(ActiveRow)].value=FFTVolts #Set the Det volts value in Excel sheet.
    ws[getExcelColumn(34+7*(ReadingNumber-1))+str(ActiveRow)].value=FFTPhase #Set the Det phase value in Excel sheet.
    print("Initial FFTVolts: " + str(FFTVolts))
    print("Initial FFTPhase: " + str(FFTPhase))
    TrialAmps = ws[getExcelColumn(64)+'7'].value
    lineCurrent = TrialAmps
    TrialVolts = ws[getExcelColumn(65)+'7'].value
    lineVolts = TrialVolts
    TrialPhase = ws[getExcelColumn(66)+'7'].value
    phase = TrialPhase
    WCount = ws['BP7'].value
    VCount = ws['BR7'].value
    WSign = ws['BQ7'].value
    VSign = ws['BS7'].value
    print WCount,VCount,WSign,VSign,lineCurrent,lineVolts,phase
    updateGUI(wattBridgeGUI,ws)
    #Output to RS232 6 WB with "DV", term.=CR, wait for completion?=1
    #Close RS232 6 WB
    RS232_6_WB.write("DV\r")
    time.sleep(1)
    final = time.clock()
    print('findDialSetting time: '+ str(final-start))
def refineDialSettings(wattBridgeGUI,ws):
    '''Obtains the Dial settings from the Excel sheet and refines them and executes the setUpFFT 
    and setUpFFTVoltsAndPhase functions. The GUI gets updated at the same time.'''
    global WCount, VCount, WSign, VSign, lineCurrent, lineVolts, phase
    start = time.clock()
    time.sleep(0.5) #Delay for 0.5 seconds
    WCount = ws['BP7'].value
    VCount = ws['BR7'].value
    WSign = ws['BQ7'].value
    VSign = ws['BS7'].value
    #Close RS232 6 WB
    #Close RS232 6 WB
    #Close RS232 6 WB
    #Close RS232 6 WB
    #Close RS232 6 WB
    #Close RS232 6 WB
    #Close RS232 6 WB
    RS232_6_WB.write("DD\r") #Output to RS232 6 WB with "DD", term.=CR, wait for completion?=1
    time.sleep(1)
    RS232_6_WB.write("W"+str(WCount)+"\r") #Output to RS232 6 WB with "W" , WCount, term.=CR, wait for completion?=1
    time.sleep(1)
    RS232_6_WB.write("V"+str(VCount)+"\r") #Output to RS232 6 WB with "V" , VCount, term.=CR, wait for completion?=1
    time.sleep(1)
    RS232_6_WB.write(str(WSign)+"\r") #Output to RS232 6 WB with WSign, term.=CR, wait for completion?=1
    time.sleep(1)
    RS232_6_WB.write(str(VSign)+"\r") #Output to RS232 6 WB with VSign, term.=CR, wait for completion?=1
    time.sleep(1)
    RS232_6_WB.write("A33\r") #Output to RS232 6 WB with "A33"(3), term.=CR, wait for completion?=1
    time.sleep(1)
    RS232_6_WB.write("B33\r") #Output to RS232 6 WB with "B33"(3), term.=CR, wait for completion?=1
    time.sleep(1)
    WattsDial = float(WCount)/1024
    VarsDial = float(VCount)/1024
    ws['AC'+str(ActiveRow)].value = WattsDial
    ws['AD'+str(ActiveRow)].value = WSign
    ws['AE'+str(ActiveRow)].value = VarsDial
    ws['AF'+str(ActiveRow)].value = VSign
    ws['AM7'].value = "Max"
    time.sleep(3) #Delay for 0.5 seconds
    setUpFFTVoltsAndPhase(ws) #Execute FFT Volts & Phase
    WattsDial = float(WCount)/1024
    VarsDial = float(VCount)/1024
    RS232_6_WB.write("W"+str(WCount)+"\r") #Output to RS232 6 WB with "W" , WCount), term.=CR, wait for completion?=1
    time.sleep(1)
    RS232_6_WB.write("V"+str(VCount)+"\r") #Output to RS232 6 WB with "V" , VCount, term.=CR, wait for completion?=1
    time.sleep(1)
    ws['AC'+str(ActiveRow)].value = WattsDial
    ws['AD'+str(ActiveRow)].value = WSign
    ws['AE'+str(ActiveRow)].value = VarsDial
    ws['AF'+str(ActiveRow)].value = VSign
    start_2 = time.clock()
    ws[getExcelColumn(33+7*(ReadingNumber-1))+str(ActiveRow)].value=FFTVolts #Set the Det volts value in Excel sheet.
    final_2 = time.clock()
    print("Set value in excel cell time: " + str(final_2-start_2))
    ws[getExcelColumn(34+7*(ReadingNumber-1))+str(ActiveRow)].value=FFTPhase #Set the Det phase value in Excel sheet.
    start_2 = time.clock()
    TrialAmps = ws[getExcelColumn(64)+'7'].value
    final_2 = time.clock()
    print("Get value in excel cell time: " + str(final_2-start_2))
    lineCurrent = TrialAmps
    TrialVolts = ws[getExcelColumn(65)+'7'].value
    lineVolts = TrialVolts
    TrialPhase = ws[getExcelColumn(66)+'7'].value
    phase = TrialPhase
    WCount = ws['BP7'].value
    VCount = ws['BR7'].value
    WSign = ws['BQ7'].value
    VSign = ws['BS7'].value
    print WCount,VCount,WSign,VSign,lineCurrent,lineVolts,phase 
    updateGUI(wattBridgeGUI,ws)
    final = time.clock()
    print('refineDialSettings: '+ str(final-start))
def loadDialSettings(ws):
    '''Loads all of the Dial settings from the Excel sheet.'''
    start = time.clock()
    #Close RS232 6 WB
    RS232_6_WB.write("DV\r") #Output to RS232 6 WB with "DV", term.=CR, wait for completion?=1
    time.sleep(1)
    WCount = ws['BP7'].value
    VCount = ws['BR7'].value
    WSign = ws['BQ7'].value
    VSign = ws['BS7'].value
    #Close RS232 6 WB
    #Close RS232 6 WB
    #Close RS232 6 WB
    #Close RS232 6 WB
    #Close RS232 6 WB
    #Close RS232 6 WB
    RS232_6_WB.write("W"+str(WCount)+"\r") #Output to RS232 6 WB with "W" , WCount, term.=CR, wait for completion?=1
    time.sleep(1)
    RS232_6_WB.write("V"+str(VCount)+"\r") #Output to RS232 6 WB with "V" , VCount, term.=CR, wait for completion?=1
    time.sleep(1)
    RS232_6_WB.write(str(WSign)+"\r") #Output to RS232 6 WB with WSign, term.=CR, wait for completion?=1
    time.sleep(1)
    RS232_6_WB.write(str(VSign)+"\r") #Output to RS232 6 WB with VSign, term.=CR, wait for completion?=1
    time.sleep(1)
    RS232_6_WB.write("A33\r") #Output to RS232 6 WB with "A33"(3), term.=CR, wait for completion?=1
    time.sleep(1)
    RS232_6_WB.write("B33\r") #Output to RS232 6 WB with "B33"(3), term.=CR, wait for completion?=1
    time.sleep(1)
    WattsDial = float(WCount)/1024
    VarsDial = float(VCount)/1024
    ws['AC'+str(ActiveRow)].value = WattsDial
    ws['AD'+str(ActiveRow)].value = WSign
    ws['AE'+str(ActiveRow)].value = VarsDial
    ws['AF'+str(ActiveRow)].value = VSign
    final = time.clock()
    print('LoadDialSettings: '+ str(final-start))
def readRadian2(ReadingsLoop,wsRS31Data):
    '''Obtain the metric values from the RD31 beloning to the phase inputs. These are added into the RD31Data sheet in
    the Excel file.'''
    global input1,input2,input3,inputTotal
    start = time.clock()
    input1=[] #Clear input 1
    input2=[] #Clear input 2
    input3=[] #Clear input 3
    inputTotal=[] #Clear Input Total
    for ReadRadLoop in range(7):
        rd31.port.open()
        data = rd31._get_metric(ReadRadLoop) #Call RD Get All Instant Data with Prog Radian ID,ReadRad loop,Rad Ph A,Rad Ph B,Rad Ph C,Rad Ph Neutral,Rad Ph Net
        rd31.port.close()
        #rd31.port.open()
        #status = rd31.ask(0x20,0,"") #Call RD Get Error Message with Prog Radian ID,RD 31 Error Message
        #print("Status: "+str(status))
        #rd31.port.close()
        input1.append(data[0]) #Append to Input 1 from Rad Ph A
        input2.append(data[1]) #Append to Input 2 from Rad Ph B
        input3.append(data[2]) #Append to Input 3 from Rad Ph C
        inputTotal.append(data[4]) #Append to Input Total from Rad Ph Net
    colOffset = -26+(ReadingsLoop*28)
    for WriteRadToExcel in range (7):
        #Calculate Input 1 Cell with row=Active Row offset=col offset loop=write Rad to exel
        RadData2Write = input1[WriteRadToExcel] #Calculate Rad data 2 write with vector=Input 1 loop=write Rad to exel
        wsRS31Data[getExcelColumn(colOffset+WriteRadToExcel)+str(ActiveRow)].value = RadData2Write #Set remote Excel link RD31 Data item Input 1 Cell to Rad data 2 write
        #Calculate Input 2 Cell with row=Active Row offset=col offset loop=write Rad to exel
        RadData2Write = input2[WriteRadToExcel] #Calculate Rad data 2 write with vector=Input 2 loop=write Rad to exel
        wsRS31Data[getExcelColumn(colOffset+WriteRadToExcel+7)+str(ActiveRow)].value = RadData2Write #Set remote Excel link RD31 Data item Input 2 Cell to Rad data 2 write
        #Calculate Input 3 Cell with row=Active Row offset=col offset loop=write Rad to exel
        RadData2Write = input3[WriteRadToExcel] #Calculate Rad data 2 write with vector=Input 3 loop=write Rad to exel
        wsRS31Data[getExcelColumn(colOffset+WriteRadToExcel+14)+str(ActiveRow)].value = RadData2Write #Set remote Excel link RD31 Data item Input 3 Cell to Rad data 2 write
        #Calculate Total Cell with row=Active Row offset=col offset loop=write Rad to exel
        RadData2Write = inputTotal[WriteRadToExcel] #Calculate Rad data 2 write with vector=Input Total loop=write Rad to exel
        wsRS31Data[getExcelColumn(colOffset+WriteRadToExcel+21)+str(ActiveRow)].value = RadData2Write #Set remote Excel link RD31 Data item Total Cell to Rad data 2 write
    final = time.clock()
    print('readRadian2 time: '+ str(final-start))
def pasteResults(ws):
    '''Obtains all of the calculated values such as MeanVolts etc from
    the Excel sheet and then places them within the ActiveRow lines
    below.'''
    global ActiveRow
    start = time.clock()
    ActiveRow=7
    CalculationResult = [] #Clear Calculation Results
    MeanWatts = ws['BW'+str(ActiveRow)].value #Index=0
    CalculationResult.append(MeanWatts)
    MeanVars = ws['BX'+str(ActiveRow)].value #Index=1
    CalculationResult.append(MeanVars)
    MeanVolts = ws['BY'+str(ActiveRow)].value #Index=2
    CalculationResult.append(MeanVolts)
    MeanAmps = ws['BZ'+str(ActiveRow)].value #Index=3
    CalculationResult.append(MeanAmps)
    for index in range(4, 24): #Obtain the data from Row=7 'CA7' onwards and add to list
        newColumn = chr(65+(index-4))
        data = ws['C'+newColumn+str(ActiveRow)].value
        CalculationResult.append(data)
    ActiveRow = RowNumber
    Result = CalculationResult[0]
    ws['BW'+str(ActiveRow)].value = Result
    Result = CalculationResult[1]
    ws['BX'+str(ActiveRow)].value = Result
    Result = CalculationResult[2]
    ws['BY'+str(ActiveRow)].value = Result
    Result = CalculationResult[3]
    ws['BZ'+str(ActiveRow)].value = Result
    for index in range(4, 24): #Take all of the data from list and add to the 'BWActiveRow' onwards
        Result = CalculationResult[index]
        newColumn = chr(65+(index-4))
        ws['C'+newColumn+str(ActiveRow)].value = Result
    final = time.clock()
    print('pasteResults time: '+ str(final-start))
def continueSequence(wattBridgeGUI,rowNumber,ws,wsRS31Data,wb):
    '''The core of the software. Contains all of the commands and function execution commands that performs
    all of the necessary measurements and calculations.'''
    start_whole = time.clock()
    global DCVRange,WCount,WSign,VCount,VSign,ReadingNumber,ActiveRow,RowNumber,SourceType
    global NumberOfReadings,Finished,ReadingNumber,ACVoltsRms,UncalFreqy,FFTVolts,FFTPhase,RDPhase
    RowNumber=int(rowNumber)
    wattBridgeGUI.WattBridgeEventsLog.AppendText("Initiating Radian \n") #Update event log.
    textToVoice(wattBridgeGUI,'Initiating Radian')
    initialiseRadian() #Run Initialise Radian function. Must add later
    time.sleep(1) #Delay for 1 second
    Finished = 0 #End of process?
    while(Finished==0):
        updateGUI(wattBridgeGUI,ws) #Reupdate variables shown in main GUI.
        WCount=0 #Clear all W and V variables.
        WSign=0
        VCount=0
        VSign=0
        if wattBridgeGUI.SetDMMRangeRefVolts.GetCurrentSelection()==0: #Less than 0.7 V rms
            DCVRange=1
        elif wattBridgeGUI.SetDMMRangeRefVolts.GetCurrentSelection()==1: #Less than 7.0 V rms
            DCVRange=10
        wattBridgeGUI.WattBridgeEventsLog.AppendText("Reading: _ \n") #Update event log.
        ReadingNumber=1
        wattBridgeGUI.WattBridgeEventsLog.AppendText("Applying Power \n") #Update event log.
        textToVoice(wattBridgeGUI,'Applying power')
        ActiveRow=RowNumber
        Phase123Cell = ws['P'+str(ActiveRow)].value #Obtain phase value from Excel sheet
        if Phase123Cell==123 or Phase123Cell==0:
            RDPhase=0
        else:
            RDPhase=Phase123Cell
        WattsOrVarsCell = str(ws['M'+str(ActiveRow)].value)#Obtain watts/vars value from Excel sheet
        if WattsOrVarsCell[0]=="v": #If it is vars
            rd31.port.open()
            if RDPhase==1:
                rd31.set_port_pulse_output(2,1,0x00) #Call Set Radian output pulse with Prog Radian ID,1,1,RD phase
            elif RDPhase==2:
                rd31.set_port_pulse_output(2,1,0x01) #Call Set Radian output pulse with Prog Radian ID,1,1,RD phase
            elif RDPhase==3:
                rd31.set_port_pulse_output(2,1,0x02) #Call Set Radian output pulse with Prog Radian ID,1,1,RD phase
            elif RDPhase==0:
                rd31.set_port_pulse_output(2,1,0x03) #Call Set Radian output pulse with Prog Radian ID,1,1,RD phase
            rd31.port.close() #always close port after performing a command on RD31
            #rd31.port.open()
            #status = rd31.ask(0x20,0,"") #Call RD Get Error Message with Prog Radian ID,RD 31 Error Message
            #print("Status: "+str(status))
            #rd31.port.close() #always close port after performing a command on RD31
            print("WattsOrVars: vars")
        elif WattsOrVarsCell[0]=="w": #If it is watts
            print("RDPhase: "+ str(RDPhase))
            rd31.port.open()
            if RDPhase==1:
                rd31.set_port_pulse_output(2,0,0x00) #Call Set Radian output pulse with Prog Radian ID,1,0,RD phase
            elif RDPhase==2:
                rd31.set_port_pulse_output(2,0,0x01) #Call Set Radian output pulse with Prog Radian ID,1,0,RD phase
            elif RDPhase==3:
                rd31.set_port_pulse_output(2,0,0x02) #Call Set Radian output pulse with Prog Radian ID,1,0,RD phase
            elif RDPhase==0:
                rd31.set_port_pulse_output(2,0,0x03) #Call Set Radian output pulse with Prog Radian ID,1,0,RD phase
            rd31.port.close() #always close port after performing a command on RD31
            #rd31.port.open()
            #status = rd31.ask(0x20,0,"") #Call RD Get Error Message with Prog Radian ID,RD 31 Error Message
            #print("Status: "+str(status))
            #rd31.port.close() #always close port after performing a command on RD31
            print("WattsOrVars: watt")
        if wattBridgeGUI.ShuntVoltsTest.GetValue()==True: #If user has checked Shunt Volts Test
            #Output to HP3488A_V with "CRESET 4", term.=LF
            time.sleep(1)
            #Output to HP3488A_V with "CLOSE 404", term.=LF
            time.sleep(1)
            #eventsLog.AppendText("Shunt Volts Test on \n")
        dateTime = str(time.asctime())
        ws['AA'+str(ActiveRow)].value=dateTime #Set the time and date in Excel sheet.
        ws['AP3'].value=RowNumber #Set the Row Number in Excel sheet.
        setPower(ws) #Execute Set Power function.
        print("DividerRange: "+str(DividerRange))
        print("Shunt: "+str(Shunt))
        print("CTRatio: "+str(CTRatio))
        print("HEGFreq: "+str(HEGFreq))
        SourceType = ws['B'+str(ActiveRow)].value #Get the Source Type from Excel sheet.
        print("SourceType: "+str(SourceType))
        if SourceType=="FLUKE" or SourceType=="FLUHIGH":
            print("FLUKE/FLUHIGH")
            powerFluke(wattBridgeGUI,ws) #Execute powerFluke function.
        elif SourceType=="CH":
            print("CH5500")
            powerCH5500(ws) #Execute CH5500 function
        elif SourceType=="HEG":
            #Execute PL10 !!!Not needed functon!!!
            print("PL10")
        else:
            print("SourceType selected in Excel file doesnt exist.")
        ActiveRow=7 #Set ActiveRow to 7.
        wattBridgeGUI.WattBridgeEventsLog.AppendText("Finding Dial Settings \n") #Update event log.
        textToVoice(wattBridgeGUI,'Finding Dial Settings')
        findDialSettings(wattBridgeGUI,ws) #Execute Find Dial Settings
        refineDialSettings(wattBridgeGUI,ws) #Execute Refine Dial Settings
        ActiveRow=RowNumber
        loadDialSettings(ws) #Execute Load Dial Settings
        NumberOfReadings = ws['K'+str(ActiveRow)].value #Get the Number of Readings value from Excel sheet.
        print("NumberOfReadings: " + str(NumberOfReadings))
        if wattBridgeGUI.ShuntVoltsTest.GetValue()==True: #If user has checked Shunt Volts Test
            #Execute Test !!!Not needed functon!!!
            print("Test")
        HP3478A_V.write('F4RAN5Z1') #Output to HP3478A_V with "F4RAN5Z1", term.=LF
        time.sleep(3) #Delay for 3 seconds
        HP3478A_V.write('T3') #Output to HP3478A_V with "T3", term.=LF
        input = HP3478A_V.read() #Enter from HP3478A_V up to 256 bytes, stop on EOS=LF
        ws['CR'+str(ActiveRow)].value = float(input) #Set remote Excel link item Temperature Cell to HP3478A_V
        wsRS31Data['A'+str(ActiveRow)].value=ActiveRow #Set Row number cell in "RD31 Data" sheet to Active Row value
        for ReadingsLoop in range(int(NumberOfReadings)):
            start = time.clock()
            ReadingNumber=ReadingsLoop+1
            start_1 = time.clock()
            wattBridgeGUI.WattBridgeEventsLog.AppendText("Doing Reading "+str(ReadingNumber)+" \n") #Update event log.
            updateGUI(wattBridgeGUI,ws) #Reupdate variables shown in main GUI
            final_1 = time.clock()
            print('GUI Single Change: ' + str(final_1-start_1))
            ActiveRow=RowNumber
            time.sleep(0.5) #Delay for 0.5 seconds
            RS232_6_WB.write('DV\r') #Output to RS232 6 WB with "DV", term.=CR, wait for completion?=1
            time.sleep(1)
            #Close RS232 6 WB
            RS232_6_WB.write('A01\r') #Output to RS232 6 WB with "A01", term.=CR, wait for completion?=1
            time.sleep(1)
            #Close RS232 6 WB
            RS232_6_WB.write('B01\r') #Output to RS232 6 WB with "B01", term.=CR, wait for completion?=1
            time.sleep(1)
            #Close RS232 6 WB
            time.sleep(0.5) #Delay for 0.5 seconds
            Ch1GateTimeCell = ws['J'+str(ActiveRow)].value #Get Ch1 gate time cell value from Excel file.
            if Ch1GateTimeCell>0.1:
                if wattBridgeGUI.SelectCounter.GetCurrentSelection()==0:
                    Ag53230A_V.write(':SENS:FREQ:GATE:TIME ' + str(Ch1GateTimeCell)) #Output to Ag53230A_V with ":SENS:FREQ:GATE:TIME " , Excel link, term.=LF
                    Ag53230A_V.write(":SENS:FUNC 'FREQ 1'") #Output to Ag53230A_V with ":SENS:FUNC 'FREQ 1'", term.=LF
                    Ag53230A_V.write(":INIT") #Output to Ag53230A_V with ":INIT", term.=LF
                    print("Counter chosen is 53230A")
                elif wattBridgeGUI.SelectCounter.GetCurrentSelection()==1:
                    #Output to HP3131A_V with ":sens:freq:arm:stop:tim " , Excel link, term.=LF
                    #Output to HP3131A_V with ":func 'freq 1'", term.=LF
                    #Output to HP3131A_V with "init", term.=LF
                    print("Counter chosen is 3131A")
            wattBridgeGUI.WattBridgeEventsLog.AppendText("Collecting Swerlein Measurements \n") #Update event log.
            textToVoice(wattBridgeGUI,'Collecting swerlein measurements')
            start_2 = time.clock()
            ACVoltsRms = SwerleinFreq.run() #Obtain Ac volts rms value using Swerleins Algorithm
            final_2 = time.clock()
            print('ACVoltsRMS time: ' + str(final_2-start_2))
            ws[getExcelColumn(35+7*(ReadingNumber-1))+str(ActiveRow)].value=ACVoltsRms #Set the Ac volts rms value in Excel sheet.
            UncalFreqy = SwerleinFreq.FNFreq() #Obtain the Frequency from 3458A
            updateGUI(wattBridgeGUI,ws) #Reupdate variables shown in main GUI.
            ws['AB'+str(ActiveRow)].value=UncalFreqy #Set the exact frequency value in Excel sheet.
            setUpFFT() #Execute Set Up FFT Function
            wattBridgeGUI.WattBridgeEventsLog.AppendText("Reference Phase \n") #Update event log.
            textToVoice(wattBridgeGUI,'Reference phase')
            setUpFFTVoltsAndPhase(ws) #Execute FFT Volts & Phase function
            ws[getExcelColumn(36+7*(ReadingNumber-1))+str(ActiveRow)].value=FFTVolts #Set the FFT ref volts value in Excel sheet.
            ws[getExcelColumn(37+7*(ReadingNumber-1))+str(ActiveRow)].value=FFTPhase #Set the FFT ref phase value in Excel sheet.
            #Close RS232 6 WB
            #Close RS232 6 WB
            #Close RS232 6 WB
            RS232_6_WB.write("DD\r") #Output to RS232 6 WB with "DD", term.=CR, wait for completion?=1
            time.sleep(1)
            RS232_6_WB.write("A33\r") #Output to RS232 6 WB with "A33", term.=CR, wait for completion?=1
            time.sleep(1)
            RS232_6_WB.write("B33\r") #Output to RS232 6 WB with "B33", term.=CR, wait for completion?=1
            time.sleep(1)
            wattBridgeGUI.WattBridgeEventsLog.AppendText("Detector volts and phase \n") #Update event log.
            time.sleep(1)
            textToVoice(wattBridgeGUI,'Detector volts and phase')
            print("Detector in session!!!!!")
            setUpFFTVoltsAndPhase(ws) #Execute FFT Volts & Phase function
            ws[getExcelColumn(33+7*(ReadingNumber-1))+str(ActiveRow)].value=FFTVolts #Set the Det volts value in Excel sheet.
            ws[getExcelColumn(34+7*(ReadingNumber-1))+str(ActiveRow)].value=FFTPhase #Set the Det phase value in Excel sheet.
            wattBridgeGUI.WattBridgeEventsLog.AppendText("Read RD31 \n") #Update event log.
            textToVoice(wattBridgeGUI,'Read R,D,31')
            readRadian2(ReadingNumber,wsRS31Data) #Execute Read Radian2 function.
            wattBridgeGUI.WattBridgeEventsLog.AppendText("Counter \n") #Update event log.
            textToVoice(wattBridgeGUI,'Counter')
            start_5 = time.clock()
            GateTime = Ch1GateTimeCell
            if GateTime>0.1:
                if wattBridgeGUI.SelectCounter.GetCurrentSelection()==1:
                    #Output to HP3131A_V with "fetc?", term.=LF
                    time.sleep(0.25) #Delay for 0.25 seconds
                    #Enter from HP3131A_V up to 256 bytes, stop on EOS=LF
                    HP3131A_V="Testing"
                    ws[getExcelColumn(38+7*(ReadingNumber-1))+str(ActiveRow)].value=HP3131A_V
                elif wattBridgeGUI.SelectCounter.GetCurrentSelection()==0:
                    output = Ag53230A_V.query('fetc?') #Output to Ag53230A_V with "fetc?", term.=LF
                    time.sleep(0.25) #Delay for 0.25 seconds
                    #Enter from Ag53230A_V up to 256 bytes, stop on EOS=LF
                    ws[getExcelColumn(38+7*(ReadingNumber-1))+str(ActiveRow)].value = float(output)
                if wattBridgeGUI.CounterChannel.GetCurrentSelection()==0:
                    if wattBridgeGUI.SelectCounter.GetCurrentSelection()==1:
                        #Output to HP3131A_V with ":sens:freq:arm:stop:tim " , Excel link, term.=LF
                        #Output to HP3131A_V with ":func 'freq 2'", term.=LF
                        #Output to HP3131A_V with ":read?", term.=LF
                        time.sleep(2) #Delay for 2 seconds
                        #Enter from HP3131A_V up to 256 bytes, stop on EOS=LF
                        HP3131A_V="Testing"
                        ws[getExcelColumn(39+7*(ReadingNumber-1))+str(ActiveRow)].value=HP3131A_V
                    elif wattBridgeGUI.SelectCounter.GetCurrentSelection()==0:
                        Ag53230A_V.write(':SENS:FREQ:GATE:TIME '+str(GateTime))#Output to Ag53230A_V with ":SENS:FREQ:GATE:TIME " , Excel link, term.=LF
                        Ag53230A_V.write(":SENS:FUNC 'FREQ 2'") #Output to Ag53230A_V with ":SENS:FUNC 'FREQ 2'", term.=LF
                        output = Ag53230A_V.write(':read?') #Output to Ag53230A_V with ":read?", term.=LF
                        time.sleep(2) #Delay for 2 seconds
                        #Enter from Ag53230A_V up to 256 bytes, stop on EOS=LF
                        ws[getExcelColumn(39+7*(ReadingNumber-1))+str(ActiveRow)].value = float(output)
                else:
                    inst_metric=[]
                    if WattsOrVarsCell[0]=="v": #If it is vars
                        rd31.port.open()
                        inst_metric = rd31._get_metric(4) #Call RD Get Instantaneous Data with Prog Radian ID,0,4,RD31 Total
                        rd31.port.close()
                        #rd31.port.open()
                        #status = rd31.ask(0x20,0,"") #Call RD Get Error Message with Prog Radian ID,RD 31 Error Message
                        #print("Status: "+str(status))
                        #rd31.port.close()
                        print("WattsOrVars: vars")
                    elif WattsOrVarsCell[0]=="w": #If it is watts
                        rd31.port.open()
                        inst_metric = rd31._get_metric(2) #RD Get Instantaneous Data with Prog Radian ID,0,2,RD31 Total
                        rd31.port.close()
                        #rd31.port.open()
                        #status = rd31.ask(0x20,0,"") #RD Get Error Message with Prog Radian ID,RD 31 Error Message
                        #print("Status: "+str(status))
                        #rd31.port.close()
                        print("WattsOrVars: watt")
                    RD31Total=inst_metric[4]
                    ws[getExcelColumn(39+7*(ReadingNumber-1))+str(ActiveRow)].value=RD31Total
            final_5 = time.clock()
            print("Counter reading time: " + str(final_5-start_5))
            textToVoice(wattBridgeGUI,'Completed Reading '+str(ReadingNumber))
            time.sleep(0.2)
            textToVoice(wattBridgeGUI,'in row '+str(ActiveRow))
            final = time.clock()
            print('Reading time: '+ str(final-start))
        #Execute Read Radian all Data !!!Not needed function!!!
        if SourceType=="CH":
            CHType = ws['B7'].value #Obtain CHType
            if CHType<99: #Perhaps not needed?
                #Output to CH5500_V with "S", term.=LF
                #Output to CH5050_V with "S", term.=LF
                print("SourceType = CH")
        elif SourceType=="HEG":
            #Output to PL10A_V with ":SOUR:OPER:STOP", term.=LF
            print("SourceType = HEG")
        elif SourceType=="FLUKE" or SourceType=="FLUHIGH":
           FLUKE_V.write("OUTP:STAT OFF") #Output to FLUKE_V with "OUTP:STAT OFF", term.=LF 
           print("SourceType = FLUKE or SourceType = FLUHIGH")
        #Close RS232 6 WB
        #Close RS232 6 WB
        #Close RS232 6 WB
        RS232_6_WB.write("DV\r") #Output to RS232 6 WB with "DV", term.=CR, wait for completion?=1
        time.sleep(1)
        RS232_6_WB.write("A01\r") #Output to RS232 6 WB with "A01", term.=CR, wait for completion?=1
        time.sleep(1)
        RS232_6_WB.write("B01\r") #Output to RS232 6 WB with "B01", term.=CR, wait for completion?=1
        time.sleep(1)
        wattBridgeGUI.WattBridgeEventsLog.AppendText("Calculating Results \n")
        textToVoice(wattBridgeGUI,'Calculating Results')
        updateGUI(wattBridgeGUI,ws)
        time.sleep(2) #Delay for 2 second
        pasteResults(ws) #Execute Paste Results function
        time.sleep(2) #Delay for 2 second
        ws['S'+str(ActiveRow)].value = ws['BY'+str(ActiveRow)].value #Voltage
        ws['T'+str(ActiveRow)].value = ws['BZ'+str(ActiveRow)].value #Current
        ws['U'+str(ActiveRow)].value = ws['CA'+str(ActiveRow)].value #Non-IEC Phase Difference
        ws['V'+str(ActiveRow)].value = ws['CL'+str(ActiveRow)].value #Power/var factor
        ws['W'+str(ActiveRow)].value = ws['CG'+str(ActiveRow)].value #Error %
        ws['X'+str(ActiveRow)].value = ws['CK'+str(ActiveRow)].value #Relative Expanded Uncertainty
        ws['Y'+str(ActiveRow)].value = ws['CJ'+str(ActiveRow)].value #Coverage Factor
        ws['Z'+str(ActiveRow)].value = ws['CO'+str(ActiveRow)].value #Impulses/w or var hr
        RowNumber=RowNumber+1 #Increment to the next Row in excel sheet.
        textToVoice(wattBridgeGUI,'Completed Row '+str(ActiveRow))
        ActiveRow = RowNumber
        wb.save()
        if ws['A'+str(ActiveRow)].value==0: #Once reached end of Excel sheet.
            Finished=1
    updateGUI(wattBridgeGUI,ws)
    textToVoice(wattBridgeGUI,'Completed Collecting Data')
    wattBridgeGUI.WattBridgeEventsLog.AppendText("---Completed collecting/measuring Data sequence--- \n") #Update event log.
    emailMessage('Watt Bridge Data Collection Completed', "Completed collecting Data Set. \n")
    final_whole = time.clock()
    print("Whole sequence time: " + str(final_whole-start_whole))
def initialiseRadian():
    '''Resets the Instantaneous Data, Min Data and Max Data in the RD31.'''
    #Call RD Assign Device with 7,Prog Radian ID. This process is already done in the 'CheckConnections' function
    #in MainProgram.py.
    #Call RD Inst Reset with Prog Radian ID. Below are the commands for this.
##     rd31.port.open()
##     reset_1 = rd31.ask(0x07,0,"\0x02") #Resets Instantaneous Data
##     rd31.port.close()
##     rd31.port.open()
##     reset_2 = rd31.ask(0x07,0,"\0x04") #Resets Instantaneous Min Data
##     rd31.port.close()
##     rd31.port.open()
##     reset_3 = rd31.ask(0x07,0,"\0x08") #Resets Instantaneous Max Data
##     rd31.port.close()
    print("Radian Initialised")
def setPower(ws):
    '''Obtains the DividerRange,Shunt,CTRatio,HEGFreq variables from Excel file as well as outputting 
    various commands to the "RS232 6 WB".'''
    global DividerRange,Shunt,CTRatio,HEGFreq
    DividerRange = ws['F'+str(ActiveRow)].value #Get the Divider Range cell value from Excel sheet
    Shunt = ws['G'+str(ActiveRow)].value #Get the Shunt cell value from Excel sheet
    CTRatio = ws['H'+str(ActiveRow)].value #Get the CT ratio cell from Excel sheet
    HEGFreq = ws['I'+str(ActiveRow)].value #Get the Set frequency cell from Excel sheet
    RS232_6_WB.write("DV\r")
    time.sleep(1)
    # Open RS232 6 WB
    # Set mode of RS232 6 WB baud rate=9600, parity="N", bits=8, stop bits=1
    # Close RS232 6 WB
    # Close RS232 6 WB
    # Close RS232 6 WB
    # Close RS232 6 WB
    # Close RS232 6 WB
    # Close RS232 6 WB
    # Close RS232 6 WB
    RS232_6_WB.write("DV\r") # Output to RS232 6 WB with "DV", term.=CR, wait for completion?=1
    time.sleep(1)
    RS232_6_WB.write("W0000\r") # Output to RS232 6 WB with "W0000", term.=CR, wait for completion?=1
    time.sleep(1)
    RS232_6_WB.write("V0000\r") # Output to RS232 6 WB with "V0000", term.=CR, wait for completion?=1
    time.sleep(1)
    RS232_6_WB.write("A01\r") # Output to RS232 6 WB with "A01", term.=CR, wait for completion?=1
    time.sleep(1)
    RS232_6_WB.write("B01\r") # Output to RS232 6 WB with "B01", term.=CR, wait for completion?=1
    time.sleep(1)
    if DividerRange==60:
        RS232_6_WB.write("R060\r") #Output to RS232 6 WB with "R" , "060", term.=CR, wait for completion?=1
        time.sleep(1)
        print("DividerRange is 60")
    else:
        RS232_6_WB.write("R"+str(DividerRange)+"\r") #Output to RS232 6 WB with "R" , Divider Range, term.=CR, wait for completion?=1
        time.sleep(1)
        print("DividerRange is not 60")
    # Close RS232 6 WB
    # Close RS232 6 WB
    # Close RS232 6 WB
    RS232_6_WB.write("WP-\r") # Output to RS232 6 WB with "WP-", term.=CR, wait for completion?=1
    time.sleep(1)
    RS232_6_WB.write("VP-\r") # Output to RS232 6 WB with "VP-", term.=CR, wait for completion?=1
    time.sleep(1)
    time.sleep(0.5) #Delay for 0.5 seconds
def updateGUI(wattBridgeGUI,ws):
    '''Updates the values shown in the main GUI. This is executed
    simultaneously with the startNewSequence via threads.'''
    wattBridgeGUI.CurrentRow.SetValue(str(RowNumber)) #Show current Row in Excel file in main GUI to user.
    wattBridgeGUI.LineCurrent.SetValue(str(lineCurrent)) #Display the Line Current value
    wattBridgeGUI.LineVolts.SetValue(str(lineVolts)) #Display the Line Volts value
    wattBridgeGUI.Phase.SetValue(str(phase)) #Display the phase value
    wattBridgeGUI.WCount1.SetValue(str(WCount)) #Display the W count value
    wattBridgeGUI.VCount.SetValue(str(VCount)) #Display the V Count value
    wattBridgeGUI.ActualFrequency.SetValue(str(UncalFreqy)) #Display the Frequency value
    print("GUI updated")
def startNewSequence(wattBridgeGUI,ws,wsRS31Data,wb):
    '''startNewSequence function is executed when user presses the "Start New Sequence (from "Start Row")".
    Leads onto continueSequence function. Contains 2 threads so that the "continueSequence" and "updateGUI"
    functions are executing simultaneously for the Suser.'''
    global lineCurrent, lineVolts, phase
    lineCurrent = "Calculating"
    lineVolts = "Calculating"
    phase = "Calculating"
    rowNumber = wattBridgeGUI.StartRow.GetValue() #Row number in excel sheet.
    RowNumber = rowNumber
    updateGUI(wattBridgeGUI,ws)
    continueSequence(wattBridgeGUI,rowNumber,ws,wsRS31Data,wb)
##     t1=threading.Thread(target=updateGUI,args=(wattBridgeGUI,ws,))
##     t2=threading.Thread(target=continueSequence,args=(wattBridgeGUI,rowNumber,ws,wsRS31Data,))
##     t1.start()
##     t2.start()
def setInstruments(HP3458,Ag53230,FLUKE,RD31,HP3478,WB):
    '''Store the VISA/Port objects belonging to all of the instruments within this class.'''
    global HP3458A_V,Ag53230A_V,FLUKE_V,rd31,HP3478A_V,RS232_6_WB
    HP3458A_V=HP3458
    SwerleinFreq.setInstrument(HP3458)
    Ag53230A_V=Ag53230
    FLUKE_V=FLUKE
    rd31=RD31
    HP3478A_V=HP3478
    RS232_6_WB=WB
def getHP3458A():
    '''Obtain the 3458A instrument.'''
    return HP3458A_V
def getAg53230A():
    '''Obtain the 53230A instrument.'''
    return Ag53230A_V
def getFLUKE():
    '''Obtain the FLUKE power supply instrument.'''
    return FLUKE_V
def getrd31():
    '''Obtain the RD31 Three Phase instrument.'''
    return rd31
def getHP3478A():
    '''Obtain the 3478A instrument.'''
    return HP3478A_V
def getWB():
    '''Obtain the Watt Bridge instrument.'''
    return RS232_6_WB